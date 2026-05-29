"""Spreadsheet operations — read, write, fill Excel and CSV files.

Single entry point: spreadsheet(action, ...) with actions:
  write_csv, write_excel, fill_form
"""
from __future__ import annotations
import csv
import json
from pathlib import Path


async def spreadsheet(args: dict) -> dict:
    """Unified spreadsheet tool.

    Actions:
      write_csv   — {path, headers, rows}
      write_excel — {path, sheets: {name: {headers, rows}}}
      fill_form   — {template_path, output_path, data, mappings?}
    """
    action = args.get("action", "")

    if action == "write_csv":
        return await _write_csv(args)
    elif action == "write_excel":
        return await _write_excel(args)
    elif action == "fill_form":
        return await _fill_form(args)
    else:
        return {"success": False, "error": f"Unknown action: {action}. Use: write_csv, write_excel, fill_form"}


async def _write_csv(args: dict) -> dict:
    path = args.get("path", "")
    headers = args.get("headers", [])
    rows = args.get("rows", [])
    if not path:
        return {"success": False, "error": "path is required"}
    if not headers:
        return {"success": False, "error": "headers is required"}

    filepath = Path(path)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return {"success": True, "path": str(filepath), "rows_written": len(rows), "columns": len(headers)}


async def _write_excel(args: dict) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "error": "openpyxl not installed. Run: pip install openpyxl"}

    path = args.get("path", "")
    sheets = args.get("sheets", {})
    if not path:
        return {"success": False, "error": "path is required"}

    filepath = Path(path)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Single sheet shorthand
    if "headers" in sheets and "rows" in sheets:
        sheets = {"Sheet1": sheets}

    total_rows = 0
    for sheet_name, data in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        headers = data.get("headers", [])
        rows = data.get("rows", [])
        if headers:
            ws.append(headers)
        for row in rows:
            ws.append(row)
        total_rows += len(rows)

    wb.save(str(filepath))
    wb.close()
    return {"success": True, "path": str(filepath), "sheets": len(sheets), "rows_written": total_rows}


async def _fill_form(args: dict) -> dict:
    template_path = args.get("template_path", "")
    output_path = args.get("output_path", "")
    data = args.get("data", {})
    mappings = args.get("mappings", {})

    if not template_path or not output_path:
        return {"success": False, "error": "template_path and output_path required"}

    tpl = Path(template_path)
    if not tpl.exists():
        return {"success": False, "error": f"Template not found: {template_path}"}

    suffix = tpl.suffix.lower()

    if suffix == ".csv":
        with open(tpl, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            template_rows = list(reader)

        col_map = mappings or {h: h for h in headers}
        filled_rows = []
        for row in template_rows:
            new_row = []
            for i, val in enumerate(row):
                col_name = headers[i] if i < len(headers) else ""
                data_key = col_map.get(col_name, col_name)
                if data_key in data and (val == "" or val is None):
                    new_row.append(str(data[data_key]))
                else:
                    new_row.append(val)
            filled_rows.append(new_row)

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        with open(out, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(filled_rows)

        return {"success": True, "path": str(out), "fields_filled": len(data), "rows_written": len(filled_rows)}

    elif suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            return {"success": False, "error": "openpyxl not installed"}

        wb = openpyxl.load_workbook(str(tpl))
        col_map = mappings or {}
        cells_filled = 0

        for ws in wb.worksheets:
            header_row = [str(c.value) if c.value else "" for c in ws[1]]
            for row in ws.iter_rows(min_row=2):
                for i, cell in enumerate(row):
                    col_name = header_row[i] if i < len(header_row) else ""
                    data_key = col_map.get(col_name, col_name)
                    if data_key in data and (cell.value is None or cell.value == ""):
                        cell.value = data[data_key]
                        cells_filled += 1

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        wb.close()
        return {"success": True, "path": str(out), "cells_filled": cells_filled}

    return {"success": False, "error": f"Unsupported format: {suffix}"}
