"""Document parser — reads PDF, Excel, CSV, DOCX, JSON, text files.

Single entry point: parse_document(path)
"""
from __future__ import annotations
import csv
import json
from pathlib import Path


def _parse_csv(filepath: str) -> dict:
    rows = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= 499:
                break
    return {"type": "csv", "headers": headers, "row_count": len(rows), "rows": rows[:50], "total_rows": len(rows)}


def _parse_excel(filepath: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets_data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append([str(c) if c is not None else "" for c in row])
            if i >= 499:
                break
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        sheets_data[sheet_name] = {"headers": headers, "row_count": len(data_rows), "rows": data_rows[:50], "total_rows": len(data_rows)}
    wb.close()
    return {"type": "excel", "sheets": sheets_data, "sheet_count": len(sheets_data)}


def _parse_pdf(filepath: str) -> dict:
    try:
        import pdfplumber
        pages = []
        with pdfplumber.open(filepath) as pdf:
            for i, page in enumerate(pdf.pages[:50]):
                text = page.extract_text() or ""
                pages.append({"page": i + 1, "text": text[:3000]})
        return {"type": "pdf", "page_count": len(pages), "pages": pages}
    except ImportError:
        pass
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        pages = []
        for i, page in enumerate(reader.pages[:50]):
            text = page.extract_text() or ""
            pages.append({"page": i + 1, "text": text[:3000]})
        return {"type": "pdf", "page_count": len(pages), "pages": pages}
    except ImportError:
        return {"error": "No PDF library installed. Run: pip install pdfplumber"}


def _parse_docx(filepath: str) -> dict:
    try:
        from docx import Document
        doc = Document(filepath)
        paragraphs = [p.text for p in doc.paragraphs[:500] if p.text.strip()]
        return {"type": "docx", "paragraph_count": len(paragraphs), "paragraphs": paragraphs}
    except ImportError:
        return {"error": "python-docx not installed. Run: pip install python-docx"}


def _parse_text(filepath: str) -> dict:
    try:
        content = Path(filepath).read_text(encoding="utf-8")
    except UnicodeDecodeError:
        content = Path(filepath).read_text(encoding="latin-1")
    return {"type": "text", "line_count": len(content.splitlines()), "content": content[:10000]}


async def parse_document(args: dict) -> dict:
    """Parse a document file. params: {path}"""
    path = args.get("path", "")
    if not path:
        return {"success": False, "error": "No path provided"}

    filepath = Path(path)
    if not filepath.exists():
        return {"success": False, "error": f"File not found: {path}"}

    suffix = filepath.suffix.lower()
    try:
        if suffix == ".csv":
            result = _parse_csv(str(filepath))
        elif suffix in (".xlsx", ".xls"):
            result = _parse_excel(str(filepath))
        elif suffix == ".pdf":
            result = _parse_pdf(str(filepath))
        elif suffix == ".docx":
            result = _parse_docx(str(filepath))
        elif suffix == ".json":
            content = filepath.read_text(encoding="utf-8")
            result = {"type": "json", "data": json.loads(content[:50000])}
        else:
            result = _parse_text(str(filepath))

        result["success"] = True
        result["filepath"] = str(filepath)
        return result
    except Exception as e:
        return {"success": False, "error": f"Failed to parse {suffix}: {str(e)}"}
